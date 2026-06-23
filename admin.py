from flask import Blueprint, render_template, jsonify, request
import sqlite3, os

admin_bp = Blueprint('admin', __name__, url_prefix='/admin')
DB_PATH  = os.path.join(os.path.dirname(__file__), 'project_mgmt.db')

def db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

# ── STEP 7b: Excel Import ─────────────────────────────────────
@admin_bp.route('/api/tasks/import', methods=['POST'])
def tasks_import():
    """
    Replace all tasks for a project with rows parsed from an uploaded Excel file.
    Expected columns: WBS, TaskDesc, ResourceName, Duration, Predecessor
    Calculates Start/Finish in row order using scheduler.py logic.
    """
    import openpyxl
    from scheduler import compute_finish, get_successor_start, parse_dt
    from datetime import datetime

    client  = request.form.get('client', '')
    project = request.form.get('project', '')
    release = request.form.get('release', '')
    file    = request.files.get('file')

    if not (client and project and release):
        return jsonify({'ok': False, 'error': 'Missing project context.'}), 400
    if not file:
        return jsonify({'ok': False, 'error': 'No file uploaded.'}), 400

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return jsonify({'ok': False, 'error': f'Could not read Excel file: {e}'}), 400

    # Read header row, map column names (case-insensitive, trimmed)
    headers = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            headers[str(cell.value).strip().lower()] = idx

    required = ['wbs', 'taskdesc']
    for r in required:
        if r not in headers:
            return jsonify({'ok': False, 'error': f"Missing required column: {r}"}), 400

    def get_cell(row, name):
        idx = headers.get(name)
        if idx is None:
            return None
        val = row[idx].value if idx < len(row) else None
        return str(val).strip() if val is not None else None

    # Parse all rows first
    rows_data = []
    for row in ws.iter_rows(min_row=2):
        wbs = get_cell(row, 'wbs')
        if not wbs:
            continue  # skip blank rows
        rows_data.append({
            'wbs':         wbs,
            'taskDesc':    get_cell(row, 'taskdesc') or '',
            'resource':    get_cell(row, 'resourcename') or '',
            'duration':    get_cell(row, 'duration') or '0',
            'predecessor': get_cell(row, 'predecessor') or '',
        })

    if not rows_data:
        return jsonify({'ok': False, 'error': 'No data rows found in the Excel file.'}), 400

    conn = db()

    # Get project's CutoverStartDateTime as default start
    proj = conn.execute(
        "SELECT CutoverStartDateTime FROM ProjectInfo WHERE Client=? AND ProjectName=? AND Release=?",
        [client, project, release]
    ).fetchone()
    default_start = proj['CutoverStartDateTime'] if proj else None

    # Get NumberRange for this project
    nr = conn.execute(
        "SELECT RangeKey, LastUsedNumber FROM NumberRange WHERE Client=? AND ProjectName=? AND Release=?",
        [client, project, release]
    ).fetchone()
    range_key = nr['RangeKey'] if nr else None
    next_unique_id = (nr['LastUsedNumber'] if nr else 0)

    # REPLACE: delete all existing tasks for this project
    conn.execute(
        "DELETE FROM Tasks WHERE Client=? AND ProjectName=? AND Release=?",
        [client, project, release]
    )
    conn.commit()

    # Determine header tasks (WBS that is a prefix of another WBS)
    all_wbs = [r['wbs'] for r in rows_data]
    def is_header(wbs):
        return any(w != wbs and w.startswith(wbs + '.') for w in all_wbs)

    errors = []
    inserted_tasks = []  # list of dicts with taskId, wbs, finish (in insertion order)

    for i, r in enumerate(rows_data):
        row_num = i + 1  # 1-based row position = predecessor reference number
        wbs = r['wbs']
        task_desc = r['taskDesc']
        resource = r['resource']
        try:
            duration = float(r['duration']) if r['duration'] else 0
        except ValueError:
            duration = 0
        predecessor = r['predecessor']
        header_flag = is_header(wbs)

        start_dt_str = None
        finish_dt_str = None
        row_error = None

        if header_flag:
            # Header tasks get rolled up later; skip calc for now
            pass
        elif duration <= 0:
            # Zero-duration / milestone task — still needs a start
            pass
        else:
            # Determine start
            if predecessor:
                pred_nums = []
                for p in predecessor.split(','):
                    p = p.strip()
                    if p.isdigit():
                        pred_nums.append(int(p))
                if pred_nums:
                    latest_finish = None
                    missing = []
                    for pn in pred_nums:
                        if 1 <= pn <= len(inserted_tasks):
                            pf = inserted_tasks[pn - 1]['finish']
                            if pf:
                                if not latest_finish or pf > latest_finish:
                                    latest_finish = pf
                            else:
                                missing.append(pn)
                        else:
                            missing.append(pn)
                    if missing:
                        row_error = f"Row {row_num}: predecessor row(s) {missing} have no finish date (likely header or zero-duration)."
                    elif latest_finish:
                        if resource:
                            first_res = resource.split(',')[0].strip()
                            start_dt_str, err = get_successor_start(latest_finish, first_res, None)
                            if err:
                                row_error = f"Row {row_num}: {err}"
                        else:
                            row_error = f"Row {row_num}: has predecessor but no resource assigned — cannot determine calendar."
                else:
                    row_error = f"Row {row_num}: predecessor '{predecessor}' is not a valid row number."
            else:
                # No predecessor — use project default start
                if default_start:
                    start_dt_str = default_start
                else:
                    row_error = f"Row {row_num}: no predecessor and project has no CutoverStartDateTime set."

            # Calculate finish if we have a start
            if start_dt_str and not row_error:
                if resource:
                    finish_dt_str, err = compute_finish({
                        'ScheduleType': 'Auto',
                        'WorkCalendarOverride': '',
                        'ResourceName': resource,
                        'StartDateTime': start_dt_str,
                        'Duration': duration,
                    })
                    if err:
                        row_error = f"Row {row_num}: {err}"
                else:
                    row_error = f"Row {row_num}: has duration but no resource assigned — cannot calculate finish."

        if row_error:
            errors.append(row_error)

        next_unique_id += 1
        team_name = None  # Could be derived from Resources table later if needed

        cur = conn.execute("""
            INSERT INTO Tasks (Client, ProjectName, Release, UniqueID, WBS, TaskDesc,
                ResourceName, Duration, Predecessor, StartDateTime, FinishDateTime,
                PerComplete, ScheduleType, TeamName)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, [client, project, release, next_unique_id, wbs, task_desc,
              resource, duration, predecessor, start_dt_str, finish_dt_str,
              0, 'Auto', team_name])

        new_task_id = cur.lastrowid
        inserted_tasks.append({
            'taskId': new_task_id, 'wbs': wbs,
            'start': start_dt_str, 'finish': finish_dt_str,
            'isHeader': header_flag, 'duration': duration,
        })

    conn.commit()

    # ── Rollup header Duration + PerComplete from leaf children ──
    wbs_map = {t['wbs']: t for t in inserted_tasks}

    def get_leaf_descendants(parent_wbs):
        leaves = []
        for w, t in wbs_map.items():
            if w == parent_wbs:
                continue
            if w.startswith(parent_wbs + '.'):
                has_children = any(
                    ow != w and ow.startswith(w + '.') for ow in wbs_map
                )
                if not has_children:
                    leaves.append(w)
        return leaves

    headers_sorted = sorted(
        [t for t in inserted_tasks if t['isHeader']],
        key=lambda t: t['wbs'].count('.'), reverse=True
    )
    for h in headers_sorted:
        leaves = get_leaf_descendants(h['wbs'])
        if not leaves:
            continue
        total_dur = sum(wbs_map[l]['duration'] or 0 for l in leaves)
        starts  = [wbs_map[l]['start']  for l in leaves if wbs_map[l]['start']]
        finishes= [wbs_map[l]['finish'] for l in leaves if wbs_map[l]['finish']]
        h_start  = min(starts) if starts else None
        h_finish = max(finishes) if finishes else None
        conn.execute(
            "UPDATE Tasks SET Duration=?, StartDateTime=?, FinishDateTime=? WHERE TaskID=?",
            [total_dur, h_start, h_finish, h['taskId']]
        )
        # Update in-memory for further rollups
        h['duration'] = total_dur
        h['start'] = h_start
        h['finish'] = h_finish

    # Update NumberRange
    if range_key:
        conn.execute("UPDATE NumberRange SET LastUsedNumber=? WHERE RangeKey=?",
                     [next_unique_id, range_key])
    conn.commit()
    conn.close()

    return jsonify({
        'ok': True,
        'imported': len(inserted_tasks),
        'errors': errors,
        'errorCount': len(errors),
    })


# ── STEP 1: WorkCalendar ──────────────────────────────────────
@admin_bp.route('/api/workcalendar', methods=['GET'])
def wc_list():
    rows = db().execute("SELECT * FROM WorkCalendar ORDER BY WorkingHrsCode").fetchall()
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/workcalendar', methods=['POST'])
def wc_save():
    d = request.json
    conn = db()
    existing = conn.execute("SELECT WorkingHrsCode FROM WorkCalendar WHERE WorkingHrsCode=?", [d['WorkingHrsCode']]).fetchone()
    if existing:
        conn.execute("""UPDATE WorkCalendar SET Description=?,ShiftStart=?,ShiftEnd=?,WorkDays=?,Timezone=?,Is24x7=?
            WHERE WorkingHrsCode=?""",
            [d.get('Description'),d.get('ShiftStart'),d.get('ShiftEnd'),d.get('WorkDays'),d.get('Timezone'),d.get('Is24x7',0),d['WorkingHrsCode']])
    else:
        conn.execute("""INSERT INTO WorkCalendar(WorkingHrsCode,Description,ShiftStart,ShiftEnd,WorkDays,Timezone,Is24x7)
            VALUES(?,?,?,?,?,?,?)""",
            [d['WorkingHrsCode'],d.get('Description'),d.get('ShiftStart'),d.get('ShiftEnd'),d.get('WorkDays'),d.get('Timezone'),d.get('Is24x7',0)])
    conn.commit()
    return jsonify({'ok': True})

@admin_bp.route('/api/workcalendar/<code>', methods=['DELETE'])
def wc_delete(code):
    conn = db()
    conn.execute("DELETE FROM WorkCalendar WHERE WorkingHrsCode=?", [code])
    conn.commit()
    return jsonify({'ok': True})

# ── STEP 2: ProjectInfo ───────────────────────────────────────
@admin_bp.route('/api/projects', methods=['GET'])
def proj_list():
    rows = db().execute("SELECT * FROM ProjectInfo ORDER BY ProjectName").fetchall()
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/projects', methods=['POST'])
def proj_save():
    d = request.json
    conn = db()
    existing = conn.execute("SELECT Client FROM ProjectInfo WHERE Client=? AND ProjectName=? AND Release=?",
        [d['Client'],d['ProjectName'],d['Release']]).fetchone()
    if existing:
        conn.execute("""UPDATE ProjectInfo SET TeamName=?,CutoverStartDateTime=?,Timezone=?
            WHERE Client=? AND ProjectName=? AND Release=?""",
            [d.get('TeamName'),d.get('CutoverStartDateTime'),d.get('Timezone'),
             d['Client'],d['ProjectName'],d['Release']])
    else:
        conn.execute("""INSERT INTO ProjectInfo(Client,ProjectName,Release,TeamName,CutoverStartDateTime,Timezone)
            VALUES(?,?,?,?,?,?)""",
            [d['Client'],d['ProjectName'],d['Release'],d.get('TeamName'),d.get('CutoverStartDateTime'),d.get('Timezone')])
    conn.commit()
    return jsonify({'ok': True})

@admin_bp.route('/api/projects/<client>/<project>/<release>', methods=['DELETE'])
def proj_delete(client, project, release):
    conn = db()
    conn.execute("DELETE FROM ProjectInfo WHERE Client=? AND ProjectName=? AND Release=?", [client,project,release])
    conn.commit()
    return jsonify({'ok': True})

# ── STEP 3: ProjectUsers ──────────────────────────────────────
@admin_bp.route('/api/projectusers', methods=['GET'])
def pu_list():
    client  = request.args.get('client','')
    project = request.args.get('project','')
    release = request.args.get('release','')
    rows = db().execute("""SELECT * FROM ProjectUsers WHERE Client=? AND ProjectName=? AND Release=?
        ORDER BY UserName""", [client,project,release]).fetchall()
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/projectusers', methods=['POST'])
def pu_save():
    d = request.json
    conn = db()
    if d.get('ID'):
        conn.execute("UPDATE ProjectUsers SET UserName=?,UserEmailID=? WHERE ID=?",
            [d['UserName'],d.get('UserEmailID'),d['ID']])
    else:
        conn.execute("""INSERT INTO ProjectUsers(Client,ProjectName,Release,UserName,UserEmailID)
            VALUES(?,?,?,?,?)""",
            [d['Client'],d['ProjectName'],d['Release'],d['UserName'],d.get('UserEmailID')])
    conn.commit()
    return jsonify({'ok': True})

@admin_bp.route('/api/projectusers/<int:uid>', methods=['DELETE'])
def pu_delete(uid):
    conn = db()
    conn.execute("DELETE FROM ProjectUsers WHERE ID=?", [uid])
    conn.commit()
    return jsonify({'ok': True})

# ── STEP 4: Resources ─────────────────────────────────────────
@admin_bp.route('/api/resources', methods=['GET'])
def res_list():
    client  = request.args.get('client','')
    project = request.args.get('project','')
    release = request.args.get('release','')
    rows = db().execute("""SELECT * FROM Resources WHERE Client=? AND ProjectName=? AND Release=?
        ORDER BY ResourceName""", [client,project,release]).fetchall()
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/resources', methods=['POST'])
def res_save():
    d = request.json
    conn = db()
    if d.get('ResourceID'):
        conn.execute("UPDATE Resources SET ResourceName=?,ResourceEmail=? WHERE ResourceID=?",
            [d['ResourceName'],d.get('ResourceEmail'),d['ResourceID']])
    else:
        conn.execute("""INSERT INTO Resources(Client,ProjectName,Release,ResourceName,ResourceEmail)
            VALUES(?,?,?,?,?)""",
            [d['Client'],d['ProjectName'],d['Release'],d['ResourceName'],d.get('ResourceEmail')])
    conn.commit()
    return jsonify({'ok': True})

@admin_bp.route('/api/resources/<int:rid>', methods=['DELETE'])
def res_delete(rid):
    conn = db()
    conn.execute("DELETE FROM Resources WHERE ResourceID=?", [rid])
    conn.commit()
    return jsonify({'ok': True})

# ── STEP 5: ResourceCalendar ──────────────────────────────────
@admin_bp.route('/api/rescalendar', methods=['GET'])
def rc_list():
    client  = request.args.get('client','')
    project = request.args.get('project','')
    release = request.args.get('release','')
    # Filter by resources belonging to this project
    rows = db().execute("""
        SELECT rc.* FROM ResourceCalendar rc
        INNER JOIN Resources r ON rc.ResourceName = r.ResourceName
            AND r.Client=? AND r.ProjectName=? AND r.Release=?
        ORDER BY rc.ResourceName, rc.EffectiveFrom
    """, [client, project, release]).fetchall()
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/rescalendar', methods=['POST'])
def rc_save():
    d = request.json
    conn = db()
    if d.get('CalendarID'):
        conn.execute("""UPDATE ResourceCalendar 
            SET ResourceName=?,ResourceEmail=?,EffectiveFrom=?,EffectiveTo=?,WorkingHrsCode=?,Notes=?
            WHERE CalendarID=?""",
            [d['ResourceName'],d.get('ResourceEmail'),d['EffectiveFrom'],d['EffectiveTo'],
             d.get('WorkingHrsCode'),d.get('Notes'),d['CalendarID']])
    else:
        conn.execute("""INSERT INTO ResourceCalendar(ResourceName,ResourceEmail,EffectiveFrom,EffectiveTo,WorkingHrsCode,Notes)
            VALUES(?,?,?,?,?,?)""",
            [d['ResourceName'],d.get('ResourceEmail'),d['EffectiveFrom'],d['EffectiveTo'],
             d.get('WorkingHrsCode'),d.get('Notes')])
    conn.commit()
    return jsonify({'ok': True})

@admin_bp.route('/api/rescalendar/<int:cid>', methods=['DELETE'])
def rc_delete(cid):
    conn = db()
    conn.execute("DELETE FROM ResourceCalendar WHERE CalendarID=?", [cid])
    conn.commit()
    return jsonify({'ok': True})

# ── STEP 6: Holidays ──────────────────────────────────────────
@admin_bp.route('/api/holidays', methods=['GET'])
def hol_list():
    rows = db().execute("SELECT * FROM Holidays ORDER BY HolidayDate").fetchall()
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/holidays', methods=['POST'])
def hol_save():
    d = request.json
    conn = db()
    if d.get('HolidayID'):
        conn.execute("""UPDATE Holidays SET HolidayDate=?,Description=?,Region=?,IsRecurringYearly=?
            WHERE HolidayID=?""",
            [d['HolidayDate'],d['Description'],d.get('Region'),d.get('IsRecurringYearly',0),d['HolidayID']])
    else:
        conn.execute("""INSERT INTO Holidays(HolidayDate,Description,Region,IsRecurringYearly)
            VALUES(?,?,?,?)""",
            [d['HolidayDate'],d['Description'],d.get('Region'),d.get('IsRecurringYearly',0)])
    conn.commit()
    return jsonify({'ok': True})

@admin_bp.route('/api/holidays/<int:hid>', methods=['DELETE'])
def hol_delete(hid):
    conn = db()
    conn.execute("DELETE FROM Holidays WHERE HolidayID=?", [hid])
    conn.commit()
    return jsonify({'ok': True})

# ── SHARED: project list for dropdowns ───────────────────────
@admin_bp.route('/api/projlist', methods=['GET'])
def projlist():
    rows = db().execute("SELECT Client,ProjectName,Release FROM ProjectInfo ORDER BY ProjectName").fetchall()
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/wccodes', methods=['GET'])
def wccodes():
    rows = db().execute("SELECT WorkingHrsCode,Description FROM WorkCalendar ORDER BY WorkingHrsCode").fetchall()
    return jsonify([dict(r) for r in rows])

# ── STEP 7: Tasks ─────────────────────────────────────────────
@admin_bp.route('/api/tasks/list', methods=['GET'])
def tasks_list():
    client  = request.args.get('client','')
    project = request.args.get('project','')
    release = request.args.get('release','')
    rows = db().execute("""
        SELECT TaskID,UniqueID,WBS,TaskDesc,Duration,StartDateTime,FinishDateTime,
               PerComplete,ResourceName,TeamName,Predecessor,Successors,Notes,
               BaselineStartDateTime,BaselineFinishDateTime,
               WorkCalendarOverride,ScheduleType
        FROM Tasks WHERE Client=? AND ProjectName=? AND Release=?
        ORDER BY TaskID
    """, [client,project,release]).fetchall()

    wbs_list = [str(r['WBS']) for r in rows]
    tasks = []
    for i, r in enumerate(rows):
        wbs    = str(r['WBS']) if r['WBS'] else ''
        is_hdr = any(w != wbs and w.startswith(wbs+'.') for w in wbs_list)
        tasks.append({
            'taskId':              r['TaskID'],
            'uniqueId':            r['UniqueID'],
            'rowNum':              i + 1,
            'wbs':                 wbs,
            'depth':               wbs.count('.'),
            'isHeader':            is_hdr,
            'taskDesc':            r['TaskDesc'] or '',
            'duration':            r['Duration'] or 0,
            'start':               r['StartDateTime'] or '',
            'finish':              r['FinishDateTime'] or '',
            'pctComplete':         r['PerComplete'] or 0,
            'resource':            r['ResourceName'] or '',
            'team':                r['TeamName'] or '',
            'predecessor':         r['Predecessor'] or '',
            'successors':          r['Successors'] or '',
            'notes':               r['Notes'] or '',
            'wcOverride':          r['WorkCalendarOverride'] or '',
            'scheduleType':        r['ScheduleType'] or 'Auto',
        })
    return jsonify(tasks)

@admin_bp.route('/api/tasks/save', methods=['POST'])
def tasks_save():
    """Save a single task (insert or update)."""
    from scheduler import compute_finish, validate_task_calendar
    d = request.json

    # Only validate/recalculate when user changed scheduling-relevant fields
    sched_fields = {'start','finish','duration','resource','wcOverride','scheduleType'}
    changed = set(d.get('changedFields', []))
    sched_changed = bool(changed & sched_fields)

    has_start    = bool(d.get('start'))
    has_resource = bool(d.get('resource') or d.get('wcOverride'))
    has_duration = int(d.get('duration') or 0) > 0

    if sched_changed and d.get('scheduleType') != 'Manual' and has_start and has_resource:
        # Validate full date range when finish present
        if d.get('finish'):
            err = validate_task_calendar({
                'ScheduleType':         d.get('scheduleType','Auto'),
                'WorkCalendarOverride': d.get('wcOverride',''),
                'ResourceName':         d.get('resource',''),
                'StartDateTime':        d.get('start',''),
                'FinishDateTime':       d.get('finish',''),
            })
            if err:
                return jsonify({'ok': False, 'error': err}), 400

        # Recalculate finish if duration present
        if has_duration and 'finish' not in changed:
            finish, err = compute_finish({
                'ScheduleType':         d.get('scheduleType','Auto'),
                'WorkCalendarOverride': d.get('wcOverride',''),
                'ResourceName':         d.get('resource',''),
                'StartDateTime':        d.get('start',''),
                'Duration':             d.get('duration',0),
            })
            if err:
                return jsonify({'ok': False, 'error': err}), 400
            d['finish'] = finish

    conn = db()
    client  = d['client']
    project = d['project']
    release = d['release']

    if d.get('taskId'):
        conn.execute("""
            UPDATE Tasks SET TaskDesc=?,Duration=?,StartDateTime=?,FinishDateTime=?,
            PerComplete=?,ResourceName=?,TeamName=?,Predecessor=?,Successors=?,Notes=?,
            WorkCalendarOverride=?,ScheduleType=?
            WHERE TaskID=?
        """, [d.get('taskDesc'),d.get('duration'),d.get('start'),d.get('finish'),
              d.get('pctComplete',0),d.get('resource'),d.get('team'),
              d.get('predecessor'),d.get('successors'),d.get('notes'),
              d.get('wcOverride') or None, d.get('scheduleType','Auto'),
              d['taskId']])
    else:
        # New task — get next UniqueID from NumberRange
        conn2 = db()
        nr = conn2.execute(
            "SELECT RangeKey,LastUsedNumber FROM NumberRange WHERE Client=? AND ProjectName=? AND Release=?",
            [client,project,release]
        ).fetchone()
        next_id = (nr['LastUsedNumber'] + 1) if nr else 1
        range_key = nr['RangeKey'] if nr else 'NEW'

        conn.execute("""
            INSERT INTO Tasks(Client,ProjectName,Release,UniqueID,WBS,TaskDesc,Duration,
            StartDateTime,FinishDateTime,PerComplete,ResourceName,TeamName,
            Predecessor,Successors,Notes,WorkCalendarOverride,ScheduleType)
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, [client,project,release,next_id,d.get('wbs',''),
              d.get('taskDesc'),d.get('duration',0),d.get('start'),d.get('finish'),
              d.get('pctComplete',0),d.get('resource'),d.get('team'),
              d.get('predecessor'),d.get('successors'),d.get('notes'),
              d.get('wcOverride') or None, d.get('scheduleType','Auto')])

        conn.execute(
            "UPDATE NumberRange SET LastUsedNumber=? WHERE RangeKey=?",
            [next_id, range_key]
        )
        task_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        d['taskId'] = task_id

    conn.commit()
    return jsonify({'ok': True, 'taskId': d.get('taskId'), 'finish': d.get('finish')})

@admin_bp.route('/api/tasks/delete', methods=['POST'])
def tasks_delete():
    """Delete one or more tasks (with children check)."""
    d       = request.json
    ids     = d.get('taskIds', [])
    force   = d.get('force', False)
    conn    = db()

    if not force:
        # Check for children
        wbs_rows = conn.execute(
            f"SELECT TaskID,WBS FROM Tasks WHERE TaskID IN ({','.join('?'*len(ids))})", ids
        ).fetchall()
        all_wbs = [r['WBS'] for r in wbs_rows]
        # Check if any other task's WBS starts with one of these
        placeholders = ','.join('?' * len(ids))
        children = conn.execute(
            f"SELECT COUNT(*) as cnt FROM Tasks WHERE TaskID NOT IN ({placeholders}) AND ({' OR '.join(['WBS LIKE ?']*len(all_wbs))})",
            ids + [str(w)+'%' for w in all_wbs]
        ).fetchone()
        if children and children['cnt'] > 0:
            return jsonify({'ok': False, 'hasChildren': True,
                'message': f"One or more selected tasks have children. Delete all {children['cnt']} child tasks too?"})

    conn.execute(f"DELETE FROM Tasks WHERE TaskID IN ({','.join('?'*len(ids))})", ids)
    conn.commit()
    return jsonify({'ok': True})

@admin_bp.route('/api/tasks/reorder', methods=['POST'])
def tasks_reorder():
    """
    Update WBS for all tasks after indent/outdent/insert/delete.
    Receives full ordered list of {taskId, wbs}.
    """
    d    = request.json
    rows = d.get('tasks', [])
    conn = db()
    for r in rows:
        conn.execute("UPDATE Tasks SET WBS=? WHERE TaskID=?", [r['wbs'], r['taskId']])
    conn.commit()
    return jsonify({'ok': True})

@admin_bp.route('/api/tasks/compute_finish', methods=['POST'])
def tasks_compute_finish():
    """Compute finish datetime for a task given start + duration + calendar."""
    from scheduler import compute_finish, get_successor_start
    d = request.json

    if d.get('predecessorFinish') and (d.get('resource') or d.get('wcOverride')):
        resource_for_lookup = d['resource'].split(',')[0].strip() if d.get('resource') else ''
        start, err = get_successor_start(
            d['predecessorFinish'],
            resource_for_lookup,
            d.get('wcOverride')
        )
        if err:
            return jsonify({'ok': False, 'error': err}), 400
        d['start'] = start

    finish, err = compute_finish({
        'ScheduleType':         d.get('scheduleType','Auto'),
        'WorkCalendarOverride': d.get('wcOverride',''),
        'ResourceName':         d.get('resource',''),
        'StartDateTime':        d.get('start',''),
        'Duration':             d.get('duration',0),
    })
    if err:
        return jsonify({'ok': False, 'error': err}), 400
    return jsonify({'ok': True, 'start': d.get('start'), 'finish': finish})
